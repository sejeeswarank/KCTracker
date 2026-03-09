"""
Export Engine — generates date-wise and range-wise exports.
Formats: PDF, Excel, TXT.
Files are temporary — generated on demand, downloaded, then deleted.
"""

import os
from config import EXPORT_FOLDER
from backend.ledger import generate_ledger
from backend.database import get_transactions_by_range


def _ensure_export_dir(username):
    user_dir = os.path.join(EXPORT_FOLDER, username)
    os.makedirs(user_dir, exist_ok=True)
    return user_dir


def _group_rows_by_date(rows):
    from collections import OrderedDict
    date_groups = OrderedDict()
    for row in rows:
        d = row["date"]
        if d not in date_groups:
            date_groups[d] = []
        date_groups[d].append(row)
    return date_groups

def _process_date_group(date, group_rows):
    debit_list, credit_list = [], []
    total_debit = total_credit = 0.0
    debit_sno = credit_sno = 1
    for row in group_rows:
        if row["debit"] > 0:
            debit_list.append({"sno": debit_sno, "name": row.get("name", ""),
                               "description": row["description"], "amount": row["debit"]})
            total_debit += row["debit"]
            debit_sno += 1
        if row["credit"] > 0:
            credit_list.append({"sno": credit_sno, "name": row.get("name", ""),
                                "description": row["description"], "amount": row["credit"]})
            total_credit += row["credit"]
            credit_sno += 1
    last_balance = group_rows[-1].get("balance")
    closing_balance = last_balance if last_balance is not None else (total_credit - total_debit)

    # Per-bank closing balances
    bank_last = {}
    for row in group_rows:
        if row.get("balance") is not None:
            bank = (row.get("source_bank") or "").strip() or "Unknown"
            bank_last[bank] = float(row["balance"])

    return {"date": date, "debit": debit_list, "credit": credit_list,
            "total_debit": total_debit, "total_credit": total_credit,
            "closing_balance": closing_balance,
            "bank_balances": bank_last}

def _build_range_ledgers(username, start_date, end_date):
    rows = get_transactions_by_range(username, start_date, end_date)
    if not rows:
        return []
        
    date_groups = _group_rows_by_date(rows)
    return [_process_date_group(d, g) for d, g in date_groups.items()]


def export_day_ledger(username, date, fmt):
    ledger = generate_ledger(username, date)
    return _generate_file(username, [ledger], fmt, f"statement_{date}")


def export_range_ledger(username, start_date, end_date, fmt):
    ledgers = _build_range_ledgers(username, start_date, end_date)
    label = f"statement_{start_date}_to_{end_date}"
    return _generate_file(username, ledgers, fmt, label)


def _generate_file(username, ledgers, fmt, label):
    if fmt == "pdf":
        return _generate_pdf(username, ledgers, label)
    if fmt == "excel":
        return _generate_excel(username, ledgers, label)
    if fmt == "txt":
        return _generate_txt(username, ledgers, label)
    raise ValueError(f"Unsupported export format: {fmt}")


def _generate_txt(username, ledgers, label):
    user_dir = _ensure_export_dir(username)
    filepath = os.path.join(user_dir, f"{label}.txt")
    lines = []
    for ledger in ledgers:
        lines += ["=" * 71, f"  Date: {ledger['date']}".center(71), "=" * 71, ""]
        lines += ["-" * 71, "DEBIT".center(71), "-" * 71,
                  f"{'S.No':<5} | {'Name':<15} | {'Description':<30} | {'Amount':>10}", "-" * 71]
        for txn in ledger["debit"]:
            lines.append(f"{txn['sno']:<5} | {str(txn['name'])[:15]:<15} | "
                         f"{str(txn['description'])[:30]:<30} | {txn['amount']:>10.2f}")
        lines += [f"{'Total Debit Amount =':>57} {ledger['total_debit']:>10.2f}", ""]
        lines += ["-" * 71, "CREDIT".center(71), "-" * 71,
                  f"{'S.No':<5} | {'Name':<15} | {'Description':<30} | {'Amount':>10}", "-" * 71]
        for txn in ledger["credit"]:
            lines.append(f"{txn['sno']:<5} | {str(txn['name'])[:15]:<15} | "
                         f"{str(txn['description'])[:30]:<30} | {txn['amount']:>10.2f}")
        lines += [f"{'Total Credit Amount =':>57} {ledger['total_credit']:>10.2f}", ""]
        lines += ["-" * 71, f"Total Account Balance = {ledger['closing_balance']:.2f}".center(71),
                  "-" * 71, "", ""]
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return filepath


def _generate_excel(username, ledgers, label):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    user_dir = _ensure_export_dir(username)
    filepath = os.path.join(user_dir, f"{label}.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for ledger in ledgers:
        ws = wb.create_sheet(title=f"{ledger['date']}"[:31])
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Ledger - {ledger['date']}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["S.No", "Name", "Description", "Amount"]
        ws["B3"] = "DEBIT"
        ws["B3"].font = Font(bold=True, color="CC0000")
        for col, h in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=h).font = Font(bold=True)
        dr_row = 5
        for txn in ledger["debit"]:
            ws.cell(row=dr_row, column=1, value=txn["sno"])
            ws.cell(row=dr_row, column=2, value=txn["name"])
            ws.cell(row=dr_row, column=3, value=txn["description"])
            ws.cell(row=dr_row, column=4, value=txn["amount"])
            dr_row += 1
        ws.cell(row=dr_row, column=3, value="Total Debit Amount =").font = Font(bold=True)
        ws.cell(row=dr_row, column=4, value=ledger["total_debit"]).font = Font(bold=True)
        ws["G3"] = "CREDIT"
        ws["G3"].font = Font(bold=True, color="006600")
        for col, h in enumerate(headers, 6):
            ws.cell(row=4, column=col, value=h).font = Font(bold=True)
        cr_row = 5
        for txn in ledger["credit"]:
            ws.cell(row=cr_row, column=6, value=txn["sno"])
            ws.cell(row=cr_row, column=7, value=txn["name"])
            ws.cell(row=cr_row, column=8, value=txn["description"])
            ws.cell(row=cr_row, column=9, value=txn["amount"])
            cr_row += 1
        ws.cell(row=cr_row, column=8, value="Total Credit Amount =").font = Font(bold=True)
        ws.cell(row=cr_row, column=9, value=ledger["total_credit"]).font = Font(bold=True)
        bal_row = max(dr_row, cr_row) + 2
        ws.merge_cells(f"C{bal_row}:G{bal_row}")
        bal_cell = ws.cell(row=bal_row, column=3,
                           value=f"Total Account Balance = {ledger['closing_balance']:.2f}")
        bal_cell.font = Font(bold=True, size=12)
        bal_cell.alignment = Alignment(horizontal="center")
        for col_letter, width in [("A", 8), ("B", 20), ("C", 45), ("D", 15),
                                   ("E", 5), ("F", 8), ("G", 20), ("H", 45), ("I", 15)]:
            ws.column_dimensions[col_letter].width = width
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# PDF Generator — matches updated web UI (ledger_details style)
# ---------------------------------------------------------------------------
class PdfRenderer:
    PAGE_W    = 210
    MARGIN    = 12
    CONTENT_W = PAGE_W - 2 * MARGIN
    GAP       = 6
    TABLE_W   = (CONTENT_W - GAP) / 2
    SNO_W  = 12
    AMT_W  = 26
    NAME_W = TABLE_W - SNO_W - AMT_W
    ROW_H  = 7
    HEAD_H = 7
    SEC_H  = 8

    HDR_BG    = (225, 230, 238)
    DR_HDR_FG = (180, 40,  30)
    CR_HDR_FG = (20,  130, 80)
    HDR_BORDER= (180, 190, 205)
    COL_HDR   = (210, 222, 235)
    COL_TXT   = (70,  90,  110)
    EVEN_ROW  = (248, 249, 251)
    DR_AMT    = (180, 40,  30)
    CR_AMT    = (20,  130, 80)
    TOT_BG    = (235, 240, 248)
    BAL_BG    = (235, 242, 252)
    BAL_FG    = (30,  90,  160)

    def __init__(self, filepath):
        from fpdf import FPDF
        self.filepath = filepath
        self.pdf = FPDF(orientation="P", unit="mm", format="A4")
        self.pdf.set_margins(self.MARGIN, self.MARGIN, self.MARGIN)
        self.pdf.set_auto_page_break(auto=True, margin=15)
        self.left_x  = self.MARGIN
        self.right_x = self.MARGIN + self.TABLE_W + self.GAP

    @staticmethod
    def clean(text):
        if text is None:
            return ""
        text = (str(text)
                .replace("\u2014", "-")
                .replace("\u2013", "-")
                .replace("\u20b9", "Rs."))
        return text.encode("latin-1", "replace").decode("latin-1")

    def render(self, ledgers):
        for ledger in ledgers:
            self._render_ledger_page(ledger)
        self.pdf.output(self.filepath)
        return self.filepath

    def _render_ledger_page(self, ledger):
        self.pdf.add_page()
        self._draw_date_title(ledger['date'])
        self._draw_main_headers()
        self._draw_column_headers()
        y_d, y_c = self._draw_table_data(ledger)
        self.pdf.set_y(max(y_d, y_c))
        self._draw_totals(ledger)
        self._draw_balance_bar(ledger)

    def _draw_date_title(self, date_str):
        self.pdf.set_font("Helvetica", "B", 13)
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.cell(0, 10, self.clean(f"Date: {date_str}"), ln=True, align="C")
        y = self.pdf.get_y()
        self.pdf.set_draw_color(*self.BAL_FG)
        self.pdf.set_line_width(0.6)
        self.pdf.line(self.MARGIN, y, self.PAGE_W - self.MARGIN, y)
        self.pdf.ln(4)

    def _draw_main_headers(self):
        y = self.pdf.get_y()
        self.pdf.set_font("Helvetica", "B", 10)
        self.pdf.set_line_width(0.3)
        self.pdf.set_draw_color(*self.HDR_BORDER)

        self.pdf.set_xy(self.left_x, y)
        self.pdf.set_fill_color(*self.HDR_BG)
        self.pdf.set_text_color(*self.DR_HDR_FG)
        self.pdf.cell(self.TABLE_W, self.SEC_H, "DEBIT", border=1, align="C", fill=True, ln=0)

        self.pdf.set_xy(self.right_x, y)
        self.pdf.set_fill_color(*self.HDR_BG)
        self.pdf.set_text_color(*self.CR_HDR_FG)
        self.pdf.cell(self.TABLE_W, self.SEC_H, "CREDIT", border=1, align="C", fill=True, ln=0)
        self.pdf.ln(self.SEC_H + 2)

    def _draw_column_headers(self):
        y = self.pdf.get_y()
        self.pdf.set_font("Helvetica", "B", 7.5)
        self.pdf.set_fill_color(*self.COL_HDR)
        self.pdf.set_text_color(*self.COL_TXT)
        self.pdf.set_draw_color(170, 185, 200)
        self.pdf.set_line_width(0.25)
        for x in [self.left_x, self.right_x]:
            self.pdf.set_xy(x, y)
            self.pdf.cell(self.SNO_W,  self.HEAD_H, "S.No",   border=1, align="C", fill=True)
            self.pdf.cell(self.NAME_W, self.HEAD_H, "Name",   border=1, align="C", fill=True)
            self.pdf.cell(self.AMT_W,  self.HEAD_H, "Amount", border=1, align="C", fill=True)
        self.pdf.ln(self.HEAD_H)

    def _draw_table_data(self, ledger):
        y_start = self.pdf.get_y()
        y_d = self._draw_data_side(self.left_x,  ledger.get("debit", []),  "debit",  self.DR_AMT,  y_start)
        y_c = self._draw_data_side(self.right_x, ledger.get("credit", []), "credit", self.CR_AMT, y_start)
        return y_d, y_c

    def _draw_data_side(self, x_pos, rows, amount_key, amt_color, y_start):
        self.pdf.set_draw_color(185, 200, 215)
        self.pdf.set_line_width(0.2)
        if not rows:
            self.pdf.set_xy(x_pos, y_start)
            self.pdf.set_fill_color(248, 249, 251)
            self.pdf.set_text_color(160, 170, 180)
            self.pdf.set_font("Helvetica", "I", 8)
            self.pdf.cell(self.TABLE_W, self.ROW_H, "No data", border=1, align="C", fill=True)
            return y_start + self.ROW_H
            
        for i, t in enumerate(rows):
            row_y = y_start + i * self.ROW_H
            even = (i % 2 == 0)
            self.pdf.set_xy(x_pos, row_y)
            self.pdf.set_fill_color(*(self.EVEN_ROW if even else (255, 255, 255)))
            self.pdf.set_font("Helvetica", "", 8)
            self.pdf.set_text_color(0, 0, 0)
            
            sno  = t.get("sno", t.get("S.No", i + 1))
            name = self.clean(t.get("name", t.get("Name", "")))[:30]
            amt  = t.get("amount", t.get(amount_key, 0))
            
            self.pdf.cell(self.SNO_W,  self.ROW_H, str(sno), border=1, align="C", fill=even)
            self.pdf.cell(self.NAME_W, self.ROW_H, name,     border=1,            fill=even)
            self.pdf.set_text_color(*amt_color)
            self.pdf.cell(self.AMT_W,  self.ROW_H, f"{amt:.2f}", border=1, align="R", fill=even)
            self.pdf.set_text_color(0, 0, 0)
        return y_start + len(rows) * self.ROW_H

    def _draw_totals(self, ledger):
        y_tot = self.pdf.get_y()
        self.pdf.set_font("Helvetica", "B", 8.5)
        self.pdf.set_fill_color(*self.TOT_BG)
        self.pdf.set_draw_color(140, 160, 180)
        self.pdf.set_line_width(0.3)

        self.pdf.set_xy(self.left_x, y_tot)
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.cell(self.SNO_W + self.NAME_W, self.ROW_H + 1, "Total Debit Amount =",
                 border=1, align="R", fill=True)
        self.pdf.set_text_color(*self.DR_AMT)
        self.pdf.cell(self.AMT_W, self.ROW_H + 1, f"{ledger['total_debit']:.2f}",
                 border=1, align="R", fill=True)

        self.pdf.set_xy(self.right_x, y_tot)
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.cell(self.SNO_W + self.NAME_W, self.ROW_H + 1, "Total Credit Amount =",
                 border=1, align="R", fill=True)
        self.pdf.set_text_color(*self.CR_AMT)
        self.pdf.cell(self.AMT_W, self.ROW_H + 1, f"{ledger['total_credit']:.2f}",
                 border=1, align="R", fill=True)

        self.pdf.set_text_color(0, 0, 0)
        self.pdf.ln(self.ROW_H + 1 + 5)

    def _draw_balance_bar(self, ledger):
        bank_balances = ledger.get("bank_balances", {})
        self.pdf.set_font("Helvetica", "B", 10)
        self.pdf.set_fill_color(*self.BAL_BG)
        self.pdf.set_draw_color(*self.BAL_FG)
        self.pdf.set_line_width(0.5)
        self.pdf.set_text_color(*self.BAL_FG)

        if bank_balances:
            banks = list(bank_balances.items())
            cell_w = self.CONTENT_W / len(banks)
            bal_y = self.pdf.get_y()
            # Draw outer border
            self.pdf.rect(self.MARGIN, bal_y, self.CONTENT_W, 12, style="FD")
            # Draw each bank label + value centered in its cell
            for i, (bank, bal) in enumerate(banks):
                x = self.MARGIN + i * cell_w
                # Draw divider between banks
                if i > 0:
                    self.pdf.set_draw_color(*self.BAL_FG)
                    self.pdf.set_line_width(0.3)
                    self.pdf.line(x, bal_y + 1, x, bal_y + 11)
                # Label
                self.pdf.set_font("Helvetica", "B", 7.5)
                self.pdf.set_text_color(*self.COL_TXT)
                self.pdf.set_xy(x, bal_y + 1.5)
                self.pdf.cell(cell_w, 4, self.clean(f"{bank.upper()} BALANCE"), align="C")
                # Value
                self.pdf.set_font("Helvetica", "B", 10)
                self.pdf.set_text_color(*self.BAL_FG)
                self.pdf.set_xy(x, bal_y + 5.5)
                self.pdf.cell(cell_w, 5, self.clean(f"= Rs.{bal:.2f}"), align="C")
            self.pdf.set_y(bal_y + 12)
        else:
            self.pdf.cell(0, 10, self.clean(f"Total Account Balance = Rs.{ledger['closing_balance']:.2f}"),
                     border=1, align="C", fill=True)

        self.pdf.set_text_color(0, 0, 0)
        self.pdf.ln(8)

def _generate_pdf(username, ledgers, label):
    """
    PDF layout matches the updated web UI:
    - Portrait A4
    - Light grey DEBIT/CREDIT headers with red/green text (no solid fill)
    - Light blue-grey column headers
    - Alternating row shading
    - Red debit amounts, green credit amounts
    - Per-bank balance bar (HDFC Balance = ... | IOB Balance = ...)
    """
    user_dir = _ensure_export_dir(username)
    filepath = os.path.join(user_dir, f"{label}.pdf")
    renderer = PdfRenderer(filepath)
    return renderer.render(ledgers)